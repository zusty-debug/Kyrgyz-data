"""
Streaming TXT -> DB importer for the mock data API.

v2 — format-aware, validated against real records:
  - TAB-separated fields
  - Records can span multiple physical lines; the buffer keeps growing
    until the tail signals completion (14-digit person_id + DOB/NULL).
  - person_id matched as `^\\d{14}$` against a single tab-field
  - DOB normalised to ISO YYYY-MM-DD, NULL when missing
  - Name = first tab-field, optionally extended for patronymic / org name
  - Region = longest 1-3 words before `область/обл/район/р-н/р н/рн`
  - City   = 1-2 words after `г.\\s*` with a negative lookahead

Flush strategy:
  * PostgreSQL  → COPY FROM STDIN (10-100× faster than INSERTs)
  * SQLite       → batched INSERT with ON CONFLICT DO NOTHING-equivalent
  * MySQL would fall back too; detection is by url prefix.

Usage:
    python -m importer.import_txt --file data/data.txt --flush auto
"""
from __future__ import annotations

import argparse
import io
import os
import re
import sys
import time
from dataclasses import dataclass
from typing import Iterator, List, Optional, Tuple

sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

_engine = None
def _get_engine():
    global _engine
    if _engine is None:
        from app.database import engine
        _engine = engine
    return _engine

def _db_kind(engine) -> str:
    return (engine.url.get_backend_name() or "").lower()

# ---------------------------------------------------------------------------
# Patterns
# ---------------------------------------------------------------------------
RE_PERSON_ID = re.compile(r"^\d{14}$")
RE_REGION = re.compile(
    r"((?:[\w\-]+\s+){1,3}(?:область|обл\.?|район|р-н|р н|рн))",
    re.IGNORECASE,
)
RE_CITY = re.compile(
    r"\bг\.?\s*([А-ЯЁ][а-яёА-ЯЁ\-]+(?:\s+(?![А-ЯЁ][а-яё\-]+\s+(?:р[\s\-.н]*н\b|район\b|\d))[А-ЯЁ][а-яёА-ЯЁ\-]+)?)",
)
PATRONYMIC_SUFFIXES = (
    "ович", "евич", "овна", "евна", "кызы", "уулу",
    "оглы", "кызы", "уулу",
)

# ---------------------------------------------------------------------------
@dataclass
class ParsedRecord:
    person_id: str
    name: Optional[str] = None
    region: Optional[str] = None
    city: Optional[str] = None
    address: Optional[str] = None
    dob: Optional[str] = None  # ISO or None
    raw: str = ""

    def is_valid(self) -> bool:
        return bool(self.person_id)

    def to_row(self):
        return (self.person_id, self.name, self.region, self.city,
                self.address, self.dob)


# ---------------------------------------------------------------------------
def _is_year(s: str) -> bool:
    return len(s) == 4 and s.isdigit() and s[:2] in ("19", "20")
def _is_month(s: str) -> bool:
    return s.isdigit() and 1 <= int(s) <= 12
def _is_day(s: str) -> bool:
    return s.isdigit() and 1 <= int(s) <= 31
def _is_full_date(y, m, d) -> bool:
    return _is_year(y) and _is_month(m) and _is_day(d)
def _looks_name_continuation(tok: str) -> bool:
    s = tok.strip().strip(",")
    if not s:
        return False
    words = s.split()
    if not words:
        return False
    sl = s.lower()
    # Reject anything that looks like a location fragment
    for g in ("обл", "район", "г.", "город", "ул.", "д.", "кв.", "с.",
              "пос.", "б/н", "р-н", "р н", "рн"):
        if g in sl:
            return False
    if words[-1].lower().endswith(PATRONYMIC_SUFFIXES):
        return True
    # Multi-word short, no geo — could be an org-name continuation.
    if 2 <= len(words) <= 3 and len(s) <= 40:
        return True
    return False


# ---------------------------------------------------------------------------
def _record_complete(toks: List[str]) -> bool:
    if not toks:
        return False
    has_pid = any(RE_PERSON_ID.match(t) for t in toks)
    if not has_pid:
        return False
    last = toks[-1].upper()
    if last == "NULL":
        return True
    if len(toks) >= 3 and _is_full_date(toks[-3], toks[-2], toks[-1]):
        return True
    return False


def iter_records(file_obj: io.TextIOBase) -> Iterator[List[str]]:
    buffer: List[str] = []
    for line in file_obj:
        tokens = [t.strip() for t in line.rstrip("\n").rstrip("\r").split("\t")]
        tokens = [t for t in tokens if t]
        if not tokens:
            continue
        buffer.extend(tokens)
        if _record_complete(buffer):
            yield buffer
            buffer = []
    if buffer:
        yield buffer


# ---------------------------------------------------------------------------
# XLSX support — openpyxl-based. Each spreadsheet row is converted to the same
# list-of-strings shape that the TXT parser expects, so the rest of the pipeline
# is unchanged.
# ---------------------------------------------------------------------------

# Mapping rules — case-insensitive substring matches against the header row.
COLUMN_ALIASES = {
    "name":      ("name", "фио", "ф.и.о", "имя"),
    "region":    ("region", "область", "обл", "регион"),
    "city":      ("city", "город", "нас.пункт", "населенный"),
    "address":   ("address", "адрес", "улица"),
    "person_id": ("person_id", "инн", "иин", "id"),
    "dob":       ("dob", "дата рождения", "дата_рождения", "др", "birth", "рожд"),
}


def iter_xlsx_records(file_path: str) -> Iterator[List[str]]:
    """
    Reads an .xlsx file. The first non-empty row is treated as the header.
    Columns are matched to canonical fields via COLUMN_ALIASES.

    Yields each data row as a list of tokens in the canonical order
        [name, region, city, address, person_id, dob_or_NULL]
    so that the existing parse_record() works as if the row had arrived as
    a tab-separated TXT line.
    """
    from datetime import date, datetime
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise RuntimeError(
            "openpyxl is not installed — add it to requirements.txt and rebuild."
        ) from e

    wb = load_workbook(file_path, read_only=True, data_only=True)
    try:
        for ws in wb.worksheets:
            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                continue

            # Find header row: first row with at least 2 non-empty strings
            header_idx = 0
            for i, r in enumerate(rows[:5]):
                if r and sum(1 for c in r if c is not None and str(c).strip()) >= 2:
                    header_idx = i
                    break
            header = [
                (str(c).strip().lower() if c is not None else "")
                for c in rows[header_idx]
            ]
            col = {}
            for i, h in enumerate(header):
                if not h:
                    continue
                for canonical, aliases in COLUMN_ALIASES.items():
                    if any(a in h for a in aliases):
                        col.setdefault(canonical, i)

            # Data rows after the header
            for row in rows[header_idx + 1:]:
                if not row or all(c is None or str(c).strip() == "" for c in row):
                    continue

                def _cell(key):
                    if key not in col or col[key] >= len(row):
                        return ""
                    v = row[col[key]]
                    if v is None:
                        return ""
                    if isinstance(v, (datetime, date)):
                        return v.strftime("%Y-%m-%d")
                    return str(v).strip()

                # Adapt the city/region/dob into shapes the existing TXT parser
                # understands:
                #   city  -> if it's not already prefixed with "г.", prepend it
                #   region -> if it doesn't already say "обл/район", append " обл."
                #   dob   -> if it's an ISO "YYYY-MM-DD", expand to 3 separate tokens
                #             so the parser's YYYY MM DD detection works.
                city = _cell("city")
                if city and not re.match(r"^\s*г\.?\s", city):
                    city = f"г. {city}"
                region = _cell("region")
                if region and not any(k in region.lower() for k in ("обл", "район", "р-н", "р н", "рн")):
                    region = region + " обл."

                dob = _cell("dob")
                dob_tokens = []
                if dob and dob.upper() != "NULL":
                    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})$", dob)
                    if m:
                        dob_tokens = [m.group(1), m.group(2), m.group(3)]
                    else:
                        dob_tokens = [dob]
                else:
                    dob_tokens = ["NULL"]

                tokens = [
                    _cell("name"),
                    region,
                    city,
                    _cell("address"),
                    _cell("person_id"),
                    *dob_tokens,
                ]
                yield tokens
    finally:
        wb.close()


# ---------------------------------------------------------------------------
def _split_name_location(tokens: List[str]) -> Tuple[List[str], List[str]]:
    if not tokens:
        return [], []
    name = [tokens[0]]
    if len(tokens) >= 2 and _looks_name_continuation(tokens[1]):
        name.append(tokens[1])
        return name, tokens[2:]
    return name, tokens[1:]


def _extract_region(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    m = RE_REGION.search(text)
    return m.group(1).strip() if m else None


def _extract_city(text: Optional[str]) -> Optional[str]:
    if not text:
        return None
    m = RE_CITY.search(text)
    return m.group(1).strip() if m else None


def parse_record(tokens: List[str]) -> ParsedRecord:
    tokens = [t.strip() for t in tokens if t.strip()]
    rec = ParsedRecord(person_id="", raw=" | ".join(tokens))
    if not tokens:
        return rec

    pid_idx = None
    for i, t in enumerate(tokens):
        if RE_PERSON_ID.match(t):
            pid_idx = i
            rec.person_id = t
            break
    if pid_idx is None:
        return rec

    if tokens[-1].strip().upper() == "NULL":
        dob_tail_len = 1
        rec.dob = None
    elif (
        len(tokens) >= pid_idx + 4
        and _is_full_date(tokens[-3], tokens[-2], tokens[-1])
    ):
        dob_tail_len = 3
        y, m_, d = tokens[-3], tokens[-2], tokens[-1]
        rec.dob = f"{y}-{int(m_):02d}-{int(d):02d}"
    else:
        dob_tail_len = 0
        rec.dob = None

    head = tokens[:pid_idx]
    name_tokens, loc_pre = _split_name_location(head)
    post_pid_end = len(tokens) - dob_tail_len
    loc_post = tokens[pid_idx + 1:post_pid_end]

    rec.name = " ".join(name_tokens).strip() or None
    full_loc = " ".join(loc_pre + loc_post).strip()
    full_loc = re.sub(r"\s+", " ", full_loc)
    rec.address = full_loc or None
    rec.region = _extract_region(full_loc)
    rec.city = _extract_city(full_loc)
    return rec


# ---------------------------------------------------------------------------
# Flush strategies
# ---------------------------------------------------------------------------
def _flush_copy(engine, batch) -> Tuple[int, int]:
    """
    PG bulk insert. Returns (new_rows, already_present_rows).
    """
    import psycopg2
    raw_conn = engine.raw_connection()
    try:
        with raw_conn.cursor() as cur:
            cur.execute("SET statement_timeout = 0")
            cur.execute(
                """
                CREATE TEMP TABLE IF NOT EXISTS staging_import (
                    person_id VARCHAR(32),
                    name       VARCHAR(255),
                    region     VARCHAR(255),
                    city       VARCHAR(255),
                    address    VARCHAR(512),
                    dob        DATE
                ) ON COMMIT DELETE ROWS
                """
            )
            cur.execute("TRUNCATE staging_import")

            buf = io.StringIO()
            for row in batch:
                cells = []
                for v in row:
                    if v is None or v == "":
                        cells.append("\\N")
                    else:
                        s = str(v).replace("\t", " ").replace("\n", " ").replace("\r", " ")
                        cells.append(s)
                buf.write("\t".join(cells) + "\n")
            buf.seek(0)
            cur.copy_expert(
                "COPY staging_import (person_id, name, region, city, address, dob) "
                "FROM STDIN WITH (FORMAT text, NULL '\\N')",
                buf,
            )
            # Get counts: how many rows from staging are NEW vs already in people.
            cur.execute(
                """
                WITH src AS (
                  SELECT person_id, name, region, city, address, dob FROM staging_import
                ),
                inserted AS (
                  INSERT INTO people (person_id, name, region, city, address, dob)
                  SELECT * FROM src
                  ON CONFLICT (person_id) DO NOTHING
                  RETURNING 1
                )
                SELECT
                  (SELECT COUNT(*) FROM inserted) AS new_rows,
                  (SELECT COUNT(*) FROM src) - (SELECT COUNT(*) FROM inserted) AS already
                """
            )
            new_rows, already = cur.fetchone()
        raw_conn.commit()
    except Exception:
        raw_conn.rollback()
        raise
    finally:
        raw_conn.close()
    return int(new_rows), int(already)


def _flush_sqlite(engine, batch) -> int:
    """SQLite bulk insert. Uses INSERT OR IGNORE for idempotence."""
    import sqlite3
    # SQLAlchemy's sqlite engine works but for speed we drop to the raw
    # sqlite3 driver and use executemany.
    raw = engine.raw_connection()
    try:
        cur = raw.cursor()
        cur.execute("PRAGMA journal_mode=WAL")
        cur.execute("PRAGMA synchronous=OFF")
        cur.executemany(
            "INSERT OR IGNORE INTO people (person_id, name, region, city, address, dob) "
            "VALUES (?, ?, ?, ?, ?, ?)",
            batch,
        )
        raw.commit()
    finally:
        raw.close()
    return len(batch)


def _flush_insert_orm(engine, batch) -> int:
    """Generic SQLAlchemy INSERT (works for any backend)."""
    from sqlalchemy import text
    rows = [
        {"person_id": r[0], "name": r[1], "region": r[2],
         "city": r[3], "address": r[4], "dob": r[5]}
        for r in batch
    ]
    with engine.begin() as conn:
        conn.execute(
            text(
                "INSERT INTO people (person_id, name, region, city, address, dob) "
                "VALUES (:person_id, :name, :region, :city, :address, :dob)"
            ),
            rows,
        )
    return len(rows)


def _pick_flusher(engine, mode: str):
    kind = _db_kind(engine)
    if mode == "auto":
        if kind == "postgresql":
            return _flush_copy
        if kind == "sqlite":
            return _flush_sqlite
        return _flush_insert_orm
    if mode == "copy":
        return _flush_copy
    if mode == "sqlite":
        return _flush_sqlite
    if mode == "insert":
        return _flush_insert_orm
    raise ValueError("unknown flush mode: " + mode)


# ---------------------------------------------------------------------------
def run_import(
    file_path: str,
    batch_size: int = 5000,
    flush_mode: str = "auto",
    log_every: int = 5000,
    malformed_log: str = "importer/malformed.log",
):
    eng = _get_engine()
    flush = _pick_flusher(eng, flush_mode)
    print(f"DB backend: {_db_kind(eng)}  flush: {flush.__name__}  "
          f"file: {file_path}  batch: {batch_size}")
    start = time.time()
    imported = 0
    already = 0
    skipped = 0
    batch = []

    os.makedirs(os.path.dirname(malformed_log) or ".", exist_ok=True)
    bad = open(malformed_log, "w", encoding="utf-8")
    try:
        ext = os.path.splitext(file_path)[1].lower()
        if ext in (".xlsx", ".xls"):
            token_iter = iter_xlsx_records(file_path)
            _close_iter = None
        else:
            fh = io.open(file_path, "r", encoding="utf-8", errors="replace")
            token_iter = iter_records(fh)
            _close_iter = fh

        for tokens in token_iter:
            rec = parse_record(tokens)
            if not rec.is_valid():
                skipped += 1
                bad.write(rec.raw + "\n---\n")
                continue
            batch.append(rec.to_row())
            if len(batch) >= batch_size:
                n_new, n_already = flush(eng, batch)
                imported += n_new
                already += n_already
                batch.clear()
                processed = imported + already
                if processed // log_every > (processed - batch_size) // log_every:
                    elapsed = time.time() - start
                    rate = processed / max(elapsed, 0.001)
                    print(f"  new={imported:,}  already={already:,}  "
                          f"seen={processed:,}  "
                          f"malformed={skipped:,}  rate={rate:,.0f}/s")
        if batch:
            n_new, n_already = flush(eng, batch)
            imported += n_new
            already += n_already
        if _close_iter is not None:
            _close_iter.close()
    finally:
        bad.close()

    elapsed = time.time() - start
    print("-" * 50)
    print(f"DONE in {elapsed:.1f}s")
    print(f"  new rows inserted : {imported:,}")
    print(f"  already in DB     : {already:,}")
    print(f"  malformed/skipped : {skipped:,}")
    print(f"  malformed log     : {malformed_log}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--file", default="data/data.txt")
    parser.add_argument("--batch", type=int, default=5000)
    parser.add_argument(
        "--flush",
        choices=["auto", "copy", "sqlite", "insert"],
        default="auto",
        help="auto picks the best strategy based on the DB backend.",
    )
    args = parser.parse_args()
    run_import(args.file, batch_size=args.batch, flush_mode=args.flush)


if __name__ == "__main__":
    main()
